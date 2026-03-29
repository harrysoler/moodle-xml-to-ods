import xml.etree.ElementTree as ET

from openpyxl import Workbook
from xmlparse import extract_questions_from


def main():
    tree = ET.parse("./test.xml")
    root = tree.getroot()

    # logging.getLogger().setLevel(logging.DEBUG)

    questions = extract_questions_from(root)

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "Tesst"

    sheet.cell(row=1, column=2, value="Pregunta 1")
    sheet.merge_cells(start_row=2,start_column=2, end_row=2, end_column=3)
    sheet.cell(row=2, column=2, value="Tipología de ")
    sheet.cell(row=3, column=2, value="Intención que tienes al evaluar con esta pregunta")
    sheet.cell(row=4, column=2, value="Competencia")
    sheet.cell(row=5, column=2, value="Nivel de dificultad")
    sheet.cell(row=6, column=2, value="Nivel en APP")
    sheet.cell(row=7, column=2, value="Contexto")

    workbook.save("test.xlsx")

if __name__ == "__main__":
    main()
